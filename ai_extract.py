import os
import json
import logging
import base64
import csv
from datetime import datetime
from dotenv import load_dotenv
from anthropic import Anthropic
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

load_dotenv()
client = Anthropic()

# Configure logging: write errors to a file, with timestamps
logging.basicConfig(
    filename="errors.log",
    level=logging.ERROR,
    format="%(asctime)s - %(message)s"
)


def clean_json_response(raw_text):
    """Strips markdown code-block formatting from an AI response and parses it as JSON."""
    if raw_text.startswith("```"):
        raw_text = raw_text.strip("`")
        raw_text = raw_text.replace("json", "", 1)
        raw_text = raw_text.strip()

    return json.loads(raw_text)


def extract_from_image(image_path):
    """Reads an invoice from an image file (.jpg or .png) using Claude's vision capability."""
    with open(image_path, "rb") as img_file:
        image_data = base64.standard_b64encode(img_file.read()).decode("utf-8")

    media_type = "image/jpeg" if image_path.lower().endswith((".jpg", ".jpeg")) else "image/png"

    response = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=500,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": image_data,
                        },
                    },
                    {
                        "type": "text",
                        "text": """Extract the following fields from this invoice image and return ONLY a valid JSON object, nothing else:
- invoice_number
- date
- due_date
- vendor
- vendor_gst_number (the GST number, if present; use null if not found)
- amount_due
- line_items (a list of objects, each with "description", "quantity", "unit_price", and "total" — use an empty list if none found)""",
                    },
                ],
            }
        ],
    )

    return clean_json_response(response.content[0].text)


def extract_invoice_data(pdf_path):
    """Reads one PDF invoice and returns extracted data as a dictionary."""
    reader = PdfReader(pdf_path)

    invoice_text = ""
    for page in reader.pages:
        invoice_text += page.extract_text() + "\n"

    response = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=500,
        messages=[
            {
                "role": "user",
                "content": f"""Extract the following fields from this invoice text and return ONLY a valid JSON object, nothing else:
- invoice_number
- date
- due_date
- vendor
- vendor_gst_number (the GST number, if present; use null if not found)
- amount_due
- line_items (a list of objects, each with "description", "quantity", "unit_price", and "total" — use an empty list if none found)

Invoice text:
{invoice_text}"""
            }
        ]
    )

    return clean_json_response(response.content[0].text)


def validate_invoice_data(data):
    """Checks extracted data for obvious problems. Returns a list of warning messages (empty list = no issues)."""
    warnings = []

    required_fields = ["invoice_number", "date", "vendor", "amount_due"]
    for field in required_fields:
        if not data.get(field):
            warnings.append(f"Missing or empty field: {field}")

    amount = data.get("amount_due", "")
    try:
        float(amount)
    except (ValueError, TypeError):
        warnings.append(f"amount_due doesn't look like a valid number: '{amount}'")

    return warnings


def is_duplicate(invoice_number, csv_path="extracted_invoices.csv"):
    """Checks if an invoice number already exists in the CSV file."""
    if not os.path.exists(csv_path):
        return False

    with open(csv_path, "r", newline="") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            if row["invoice_number"] == invoice_number:
                return True

    return False


def format_line_items(line_items):
    """Converts a list of line item dictionaries into one readable text block."""
    if not line_items:
        return ""

    lines = []
    for item in line_items:
        line = f"{item.get('description', '')} (qty: {item.get('quantity', '')}, price: {item.get('unit_price', '')}, total: {item.get('total', '')})"
        lines.append(line)

    return " | ".join(lines)


def save_to_csv(data, csv_path="extracted_invoices.csv"):
    """Appends one row of extracted data to the CSV file."""
    fieldnames = ["invoice_number", "date", "due_date", "vendor", "vendor_gst_number",
                  "amount_due", "line_items", "needs_review", "review_notes"]
    file_exists = os.path.exists(csv_path)

    row_data = data.copy()
    row_data["line_items"] = format_line_items(data.get("line_items", []))

    with open(csv_path, "a", newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        if not file_exists or os.path.getsize(csv_path) == 0:
            writer.writeheader()
        writer.writerow(row_data)


def export_to_excel(csv_path="extracted_invoices.csv", excel_path="extracted_invoices.xlsx"):
    """Reads the CSV data and creates a nicely formatted Excel file."""
    if not os.path.exists(csv_path):
        print("No CSV file found to export.")
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoices"

    with open(csv_path, "r", newline="") as csvfile:
        reader = csv.DictReader(csvfile)
        fieldnames = reader.fieldnames

        sheet.append(fieldnames)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in reader:
            values = [row[field] for field in fieldnames]
            sheet.append(values)

            if row.get("needs_review") == "True":
                row_number = sheet.max_row
                highlight = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                for cell in sheet[row_number]:
                    cell.fill = highlight

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    workbook.save(excel_path)
    print(f"Excel file saved: {excel_path}")


def log_run_summary(stats, log_path="run_history.csv"):
    """Appends one row summarizing this entire run to the run history log."""
    fieldnames = ["timestamp", "files_found", "processed", "skipped_duplicates",
                  "failed", "needs_review_count"]
    file_exists = os.path.exists(log_path)

    row = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "files_found": stats["files_found"],
        "processed": stats["processed"],
        "skipped_duplicates": stats["skipped_duplicates"],
        "failed": stats["failed"],
        "needs_review_count": stats["needs_review_count"],
    }

    with open(log_path, "a", newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        if not file_exists or os.path.getsize(log_path) == 0:
            writer.writeheader()
        writer.writerow(row)


def print_run_summary(stats):
    """Prints a clear end-of-run summary to the terminal."""
    print("\n--- Run Summary ---")
    print(f"Files found:         {stats['files_found']}")
    print(f"Successfully saved:  {stats['processed']}")
    print(f"Skipped (duplicate): {stats['skipped_duplicates']}")
    print(f"Failed:              {stats['failed']}")
    print(f"Needs review:        {stats['needs_review_count']}")
    print("-------------------\n")


# Process every PDF or image in the "invoices" folder
invoice_folder = "invoices"

stats = {
    "files_found": 0,
    "processed": 0,
    "skipped_duplicates": 0,
    "failed": 0,
    "needs_review_count": 0,
}

for filename in os.listdir(invoice_folder):
    file_path = os.path.join(invoice_folder, filename)
    lower_name = filename.lower()

    if lower_name.endswith(".pdf"):
        is_image = False
    elif lower_name.endswith((".jpg", ".jpeg", ".png")):
        is_image = True
    else:
        continue  # skip files that aren't PDFs or supported images

    stats["files_found"] += 1
    print(f"Processing {filename}...")
    try:
        if is_image:
            data = extract_from_image(file_path)
        else:
            data = extract_invoice_data(file_path)

        if is_duplicate(data["invoice_number"]):
            print(f"  -> Skipped (already processed): {data['invoice_number']}")
            stats["skipped_duplicates"] += 1
        else:
            warnings = validate_invoice_data(data)
            data["needs_review"] = bool(warnings)
            data["review_notes"] = "; ".join(warnings)

            save_to_csv(data)
            stats["processed"] += 1
            if warnings:
                stats["needs_review_count"] += 1

            if warnings:
                print(f"  -> Saved WITH WARNINGS: {data}")
            else:
                print(f"  -> Saved: {data}")
    except Exception as e:
        error_message = f"Failed to process {filename}: {e}"
        print(f"  -> {error_message}")
        logging.error(error_message)
        stats["failed"] += 1

# After processing all invoices, generate a formatted Excel export
export_to_excel()

# Log this run's summary and print it
log_run_summary(stats)
print_run_summary(stats)